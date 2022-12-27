from flask import Flask, redirect, url_for, render_template, request, flash
from openpyxl import load_workbook
import random       # Sun Tzu Quotes and Target Assignments
import datetime     # Date and time for calendar and updates
import hashlib      # Profile URL and Team ID encoding
import os, sys      # File management

FOLDER_PATH = sys.path[0]

RULESFILE = rf"{FOLDER_PATH}\RulesOfWar.txt"
DATAFILE = rf"{FOLDER_PATH}\IndependenceAssassins.xlsx"
QUOTEFILE = rf"{FOLDER_PATH}\WarQuotes.txt"
UPDATESFILE = rf"{FOLDER_PATH}\Updates.txt"
CALENDARINPUT = rf"{FOLDER_PATH}\Calendar.xlsx"
CALENDARFILE = rf"{FOLDER_PATH}\Calendar.txt"
CLASSCOUNT = 500
ENTRYCOST = 10          # Unit is dollars


def generateRules():                                                    # Loads RULESFILE when needed
    ruleFile = open(RULESFILE,"r")
    rules = []
    for line in ruleFile.readlines():
        rules.append(line)
        rules.append("<br>")
        
    ruleFile.close()
        
    global ruleList
    ruleList = "".join(rules)
    
    return ruleList


class accounts():
    
    def __init__(self):
        pass
    
    def admin(self):
        pass
    
    def generateID(self, firstName: str, lastName:str, number: str):
        id_temp = (firstName[:1].lower()+lastName.lower())       # Player ID's will be the first letter of their first name, first 5 letters of their last name and the last 4 digits of their phone number
        id_characters = []
        id = 0
        for char in id_temp:                                        # getting ASCII code for each letter in id
            id_characters.append(ord(char))
        
        for digit in number:                                       # adding up every number in phone number
            id += int(digit)    
        
        for character in id_characters:
            id += int(character)
            
        encode_id = (firstName[:1].lower() + lastName.lower() + number[-4:])
        
        return id, encode_id
        
    def newAccount(self, firstName: str, lastName: str,password: str,number: str):              # Checks if account ID exists then creates account with entered information
        wb = load_workbook(DATAFILE)
        
        '''
            Usable columns and rows in each sheet (all are automatically filled by bot):
            Players:
                A2-A1000:Player ID's (See line 32 to see how they are generated)
                B2-B1000: First Names
                C2-C1000: Last Names
                D2-D1000: Status (Dead or Alive)
                E2-E1000: Special Titles (Space to show if anyone has done anything interesting)
                F2-F1000: Phone Numbers (Both for group chat and for player information)
                H2-H1000: URL Extensions to their profile accounts
        '''
        
        sheet = wb["Players"]
        next_row = len(sheet['A']) + 1
        
        match = False                           # match will be used to tell code whether or not found an already existing ID
        
        # checking for matches of ID's

        for row in range(2, next_row):
            if id == (sheet[f"A{row}"].value):
                
                match = True                            # if lines 52-53 finds a match to the potential new id, it will throw 409 (Conflict Error) to show this id already exists
                
        if match == True:
            return 409
        
        elif match == False:
        
            # processing Player ID into database
            id = self.generateID(firstName, lastName, number)
            
            playerID = sheet.cell(row=next_row, column = 1)
            playerID.value = id[0]
            
            # processing First Name into database
            fName = sheet.cell(row=next_row, column = 2)
            fName.value = firstName
            
            # processing Last Name into database
            lName = sheet.cell(row=next_row, column = 3)
            lName.value = lastName
            
            # processing Phone Number into database
            phoneNumber = sheet.cell(row = next_row, column = 6)
            phoneNumber.value = number
            
            # processing password into database
            passcode = sheet.cell(row = next_row, column = 7)
            passcode.value = password
            
            # setting new Player to Alive
            status = sheet.cell(row = next_row, column = 4)
            status.value = "alive"
            
            # setting number of confirmed eliminations to 0
            eliminations = sheet.cell(row = next_row, column = 5)
            eliminations.value = 0
            
            # creating user's url extension to their profile page
            url_extension = sheet.cell(row = next_row, column = 8)
            url = hashlib.sha256(id[1].encode())
            url_extension.value = url.hexdigest()
                
            
            wb.save(DATAFILE)
            
            return 200                                      # if the function is run successfully (doesn't catch a matching id), it will return 200 (OK) to allow the code to continue
        
    def validateAccount(self, name, phonenumber):    
        id = []
        
        temp_name = name.lower()
        first_initial = temp_name[:1]
        id.append(first_initial)
        
        for i, char in enumerate(temp_name):
            if char == " ":
                last_name = temp_name[i+1:]
                
        id_digits = str(phonenumber)[-4:]
        id.append(id_digits)
        
        
        potential_account_id = self.generateID(firstName = first_initial, lastName = last_name, number = phonenumber)   

        # Searching digits of ids in database for possible matchs, if there are no matches, then there must not be an account registered under that name
        database = load_workbook(DATAFILE)
        sheet = database["Players"]
        next_row = len(sheet['A']) + 1
        
        # match = False                           # match will be used to tell code whether or not found an already existing ID
        
        # checking for matches of ID's

        for row in range(2, next_row):
            if potential_account_id[0]== (sheet[f"A{row}"].value):          # checking if ASCII id matches any id in the spreadsheet
                match = True
                break
                
            else:
                match = False
                
            
        if match == False:
            return 404, " "
        
        elif match == True:
            return 403, potential_account_id[1]         # returns 403 (https error for saying something is already there which is what we want it to throw) and the id that will be encoded to create the subpage
        
    def getAccount(self, data_given: str, data_known):
        """_summary_

        Args:
            data_given (str): Type of data that is known (id, name, status, url extension)
            data_known (_type_): Data that is already known (Known name, known id, etc.)

        Returns:
            _type_: _description_
        """
        database = load_workbook(DATAFILE)['Players']
        def get_data_lists(data_given):
            columns = {
                # First Array will always be the omitted columns
                'id':([['A'],['B','C','D','E','F','G','H']]),
                'name':([['B','C'],['A','D','E','F','G','H']]),
                'status': ([['D'],['A','B','C','E','F','G','H']]),
                'url_extension': ([['H'],['A','B','C','D','E','F','G']])
            }
            
            return columns.get(data_given,404)
            
        data_columns = get_data_lists(data_given)[1]
        omitted_columns = get_data_lists(data_given)[0]
        
        data = []
        '''
            Usable columns and rows in each sheet (all are automatically filled by bot):
            Players:
                A2-A1000:Player ID's (See line 32 to see how they are generated)
                B2-B1000: First Names
                C2-C1000: Last Names
                D2-D1000: Status (Dead or Alive)
                E2-E1000: Number of Confirmed Eliminations
                F2-F1000: Phone Numbers (Both for group chat and for player information)
                G2-G1000: Password
                H2-H1000: URL Extensions to their profile accounts
        '''
        
        # Return everything on the account with the id given
        for row in range(2, len(database['A']) + 1):
            if database[f'{omitted_columns[0]}{row}'].value == str(data_known):
                for column in data_columns:
                    data.append(database[f"{column}{row}"].value)
            
            
        return data
        ''' In order data returns:
            [ID (if given url_extension), FIRST NAME, LAST NAME, STATUS (Dead or Alive), NUMBER OF CONFIRMED ELIMINATIONS, PHONE NUMBER, PASSWORD]
            or
            [FIRST NAME, LAST NAME, STATUS (Dead or Alive), NUMBER OF CONFIRMED ELIMINATIONS, PHONE NUMBER, PASSWORD, URL_EXTENSION (if given ID)]
        '''            
           
           
        
def generateQuote():        # Picks random Sun Tzu Quote from QUOTEFILE variable above
    
    # Geting Line Count
    file = open(QUOTEFILE,"r")
    for i, line in enumerate(file):
        pass
    lastLine = i

    
    # Getting random number
    quoteNumber = random.randint(0,lastLine)
    
    # matching random number to quote
    file.seek(0)                            # starting search back at the top.  Lines 98-100 push the "start" of the file to the last line.  Using file.seek(0) pushes the "start" back to line 1
    quotes = file.readlines()
    randomQuote = quotes[quoteNumber]
    
    return randomQuote
    
    file.close()
    
def cashPrize():            # Divide cash prize by current number of alive players from DATAFILE above
    wb = load_workbook(DATAFILE)
    
    '''
        Usable columns and rows in each sheet (all are automatically filled by bot):
        Players:
            A2-A1000:Player ID's (See line 32 to see how they are generated)
            B2-B1000: First Names
            C2-C1000: Last Names
            D2-D1000: Status (Dead or Alive)
            E2-E1000: Special Titles (Space to show if anyone has done anything interesting)
            F3-F1000: Phone Numbers (Both for group chat and for player information)
    '''
    sheet = wb["Players"]
    row_count = len(sheet['D'])
    
    alive_count = row_count
    
    for row in range(2,row_count):
        if sheet[f"D{row}"] != "alive":
            alive_count -= 1                        # Code will subtract 1 every time it does not find 'alive' in the Status slot
    
    cashprize = (ENTRYCOST*row_count)/alive_count
    
    return cashprize
    
def showUpdates():
    updateFile = open(UPDATESFILE,"r")
    updates = []
    for line in updateFile.readlines():
        updates.append(line)
        updates.append("<br>")
        
    updateFile.close()
        
    global updateList
    updateList = "".join(updates)
    
    return updateList
    
def newUpdate(update):      # To be added to the admin console to publish updates.
    updateFile = open(UPDATESFILE,"a")
    time = datetime.datetime.now()
    time = time.strftime("%m-%d-%Y %H:%M")
    update = (f"[{time}]: {update}")
    updateFile.write(update + "\n")
    updateFile.close()

class calendar():
    
    def __init__(self):
        self.CALFILE = CALENDARFILE
        self.CALINPUT = CALENDARINPUT
        
    def wipeFile(self):
        with open(self.CALFILE, "r+") as file:
            file.truncate(0)
            
        return 200
        
    def processCalendar(self):                                  
        '''Converting the Excel Spreadsheet of Calendar dates to text file 
        to be shown by computer on homepage'''
        calendarFile = open(self.CALFILE,"a")
        calendar = load_workbook(self.CALINPUT, data_only=True)["Sheet1"]
        calendarlist = []
        for i, row in enumerate(calendar['A']):
            event = []

            if (type(row.value) == int) or (type(row.value) == str):                # if statement to stop code once it hits an empty cell.  
                if i == 0:
                    continue
                
                print()
                if (int(row.value) > -1):
                    date = calendar[f"B{i+1}"].value
                    item = calendar[f"C{i+1}"].value
                    
                    formatted_date = str(datetime.datetime.strftime(date, "%m-%d-%Y"))
                    year = formatted_date[-4:]
                    day = formatted_date[-7:-5]
                    month = formatted_date[:-8]
                                    
                    calendarFile.write(formatted_date + " ")
                    calendarFile.write(item)
                    calendarFile.write("\n")
                    
                else:
                    continue
            else:
                break
            
        calendarFile.close()
        
    def showCalendar(self):
        
        self.wipeFile()
        self.processCalendar()
        
        calendarFile = open(self.CALFILE,"r")
        calendarList = []
        for i, row in enumerate(calendarFile.readlines()):
            calendarList.append(row)
            calendarList.append("<br>")
            calendarList.append("<br>")
            calendarList.append("<br>")
        
        return "".join(calendarList)

class teams():
    '''
        Required functions in teams:
            - Team target assignment
            - Chances of outcome (Percentages of each team winning)
                - Base on number of kills and number of people remaining compared to other team
                    - Number of eliminations increases points (Shows they actually play the game)
                    - Lower number of team members alive lowers points (Lower number of targets for other team to get)
                    - Number of revives on each team increases points (Shows monetary and teamwork advantage)
                    - Number of underdog wins (winning despite going into round with fewer numbers) multiplies score by 1.5
                - Stats will be compared to find chances of outcome
                - Stats will also be used for leaderboard ranking 
    '''
    
    def __init__(self) -> None:
        self.DATAFILE = load_workbook(DATAFILE, data_only=True)
        self.TEAMSHEET = self.DATAFILE['Teams']
        
    def generateID(self, members):
        # Team ID's will be generated by combining all names into one string, hash encyrpting the string, and converting every charcter into number and adding them together.
        
        team = "".join(members)
        encode = hashlib.sha256(team.encode())          #Encoding with sha256 encryption
        team_id = 0
        
        for char in encode.hexdigest():             # hexdigest function converts the encoded variable into a string to be iterable
            if type(char) != int:
                team_id += ord(char)         # If character is not a number (thus would be a string), find the ASCII code for the letter and add that instead of the letter.
            
            elif type(char) == int:
                team_id += char
                
        return team_id
        
    def set_teams(self, team_name: str, team_members: list):
        """Group teams together based on ids.  
        ID's will be inputed and the function will find the names of everyone as well 
        as create an ID for the team and plug all information into the 'Teams' sheet
        in the excel spreadsheet. Function will also input the team name into the spreadsheet

        Args:
            team_members (list): Enter the list of the member ids to get Player information
        """
        
        next_clear_row = len(self.TEAMSHEET['A'])+1     # Finding next available slot
        
        '''
            Notes on each necessary column:
                A Column: Team ID (Generated by generateID() function starting at line 358)
                B Column: Team Name
                C Column: Team Member 1's ID (Found with accounts().getAccount() function starting at line 172)
                D Column: Team Member 1's Status (Dead or Alive)
                E Column: Team Member 2's ID (Found with accounts().getAccount() function starting at line 172)
                F Column: Team Member 2's Status (Dead or Alive)
                G Column: Team Member 3's ID (Found with accounts().getAccount() function starting at line 172)
                H Column: Team Member 3's Status (Dead or Alive)
                I Column: Team Member 4's ID (Found with accounts().getAccount() function starting at line 172)
                J Column: Team Member 4's Status (Dead or Alive)
        '''
        
        member_names = []
        
        for i, player in enumerate(team_members):               # Uploading player data and gathering names for team ID
            # splitting first and last name
            for j, char in enumerate(player):
                if char == " ":
                    first_name = player[:j]
            
            player_account = accounts().getAccount('name', first_name)
            status = player_account[1]
            id = player_account[0]
            
            member_names.append(first_name)
            
            self.TEAMSHEET[f'{chr(67+(2*i))}{next_clear_row}'].value = id
            self.TEAMSHEET[f'{chr(68+(2*i))}{next_clear_row}'].value = status
            
            
            self.TEAMSHEET[f'B{next_clear_row}'].value = team_name
            self.TEAMSHEET[f'A{next_clear_row}'].value = self.generateID(member_names)
            
        self.DATAFILE.save(DATAFILE) 

# Everything above this line is backend management, below this line is front end:
###########################################    

app = Flask(__name__)

@app.route('/', methods = ["POST", "GET"])
def home():
    
    if request.method == "POST":
        
        if request.form["login"]:
            return redirect(url_for('login'))
    
    return render_template('home.html', 
                           ruleList = generateRules(), 
                           quote = generateQuote(), 
                           cashPerWinner = cashPrize(),
                           updates = showUpdates(),
                           events = calendar().showCalendar())

@app.route('/login/', methods = ["POST","GET"])
def login():
    
    global signin_stauts
    
    account_status = " "
    
    if request.method == "POST":
        if request.form.get("Login",False):
            name = request.form.get("username")
            number = request.form.get("phonenumber")
            
            if name == "c1f96b08fa7efdfb3732fca9db56e39a594944b2b14c5a95cce11a2e24de5b2d":
                return redirect(url_for('admin'))
            
            validAccount = accounts().validateAccount(name,number)
            if validAccount[0] == 404:
                print("No Account")
                account_status = f"There is no account tied to {name} with the phone number {number}.<br><br> Contact Gamemasters if this is a mistake or if you have forgotten your login information."
                
            
            if validAccount[0] == 403:
                signin_status = True
                print("Account found")
                url = hashlib.sha256(validAccount[1].encode())
                return redirect(url_for('profile',extension = url.hexdigest()))
        

    
    return render_template('login.html', 
                        ruleList = generateRules(),
                        events = calendar().showCalendar(),
                        account_status = account_status)

@app.route('/profile/<extension>')
def profile(extension):
    
    profile_information = accounts().getAccount(url_extension=extension)
    
    name = profile_information[1] + " " + profile_information[2]
    if profile_information[3] == "alive":
        status = "Not dead yet."
    elif profile_information[3] == "dead":
        status = "Fought hard. Died hard."
    
    eliminations = profile_information[4]
    phone_number = profile_information[5]
    
    return render_template("profile.html", extension = extension, user = name, status = status, phone_number = phone_number, eliminations = eliminations)

@app.route('/admin/c1f96b08fa7efdfb3732fca9db56e39a594944b2b14c5a95cce11a2e24de5b2d', methods = ["POST","GET"])
def admin():
    confirmation = ""
    if request.method == "POST":
        
        if request.form.get("update_submit",False):
            
            update_content = request.form.get("update_text")
            
            newUpdate(update_content)
            
            confirmation = "Update has been posted."
            
            pass
    
    return render_template('admin.html', confirm = confirmation)


if __name__ == "__main__":
    app.run(host = "0.0.0.0", debug = True)

    